from sheet_model import Sheet
from openpyxl import load_workbook

class SheetUtils:

    @staticmethod
    def _ensure_sheet_instance(sheet):
        if isinstance(sheet, Sheet):
            return sheet.get_sheet() 
        return sheet

    @staticmethod
    def _remove_duplicates_for_sheet(sheet, start_row, end_row, col_range):
        """Remove duplicate rows from the sheet"""
        sheet_obj=SheetUtils._ensure_sheet_instance(sheet)
        seen = set()
        rows_to_delete = []
        for row in sheet_obj.iter_rows(min_row=start_row, max_row=end_row, min_col=col_range[0], max_col=col_range[1]):
            row_values = tuple(cell.value for cell in row)
            if row_values in seen:
                rows_to_delete.append(row[0].row)
            else:
                seen.add(row_values)

        print(len(seen))
        for row_idx in sorted(rows_to_delete, reverse=True):
            sheet_obj.delete_rows(row_idx)
        return sheet_obj

    @staticmethod 
    def clear_row(sheet, min_row):
        for row in sheet.iter_rows(min_row=min_row):
            for cell in row:
                cell.value = None

    @staticmethod
    def _clear_rows(file, min_row):
        wb = load_workbook(file)
        for sheet in wb.worksheets:
            SheetUtils.clear_row(sheet, min_row)


    @staticmethod
    def _num_or_rows(sheet):
        sheet_obj=SheetUtils._ensure_sheet_instance(sheet)
