from openpyxl import load_workbook
import pandas as pd

# Load the workbook
wb = load_workbook("input_200rows.xlsx")

# Select the sheet
sheet = wb['ProLease Report']

# Extract data, ignoring None values
data=[]
for row in sheet.iter_rows(min_row=5, max_col=8, values_only=True):
    data.append(row)

# Convert to DataFrame
df = pd.DataFrame(data)

# Display the DataFrame
df = df.drop_duplicates().reset_index(drop=True)

# Clear existing rows (except headers)
sheet.delete_rows(5, sheet.max_row)  # Delete all rows below headers

# Write DataFrame back to sheet
for row in df.itertuples(index=False, name=None):
    sheet.append(row)

wb.save("prolease.xlsx")

