import pandas as pd
from openpyxl import load_workbook



df_in = pd.read_excel("input.xlsx", header =3)
df_tem = pd.read_excel("template_no_data.xlsx", header=2)
df_out = pd.read_excel("output_nre.xlsx", header = 2)

# print(df)

print(df_in.head())
print(df_tem.head())
print(df_out.head())
# print(df.columns)
# print(df.iat[0,1])


# print(df['LocationId'])

# wb = load_workbook("template_no_data.xlsx")
# i=0
# for sheet_name in wb.sheetnames:
    # sheet = wb[sheet_name]

    # print(sheet.title)

    # headers = [ cell.value for cell in sheet[3]]
    # print(headers)

    # for header in headers:
        # print(header)
        # for row in 
        #
        # for cell in row:
            # print(cell.value)
    # for row in sheet.iter_rows(min_row=4 ,values_only=True):
        # print(row)
    # for cell in sheet[4]:
    #     print(cell.value)
    #     cell.value=None
    #      
        # for cell in sheet[1]:
            # cell.value=None
            # print(cell) 
# wb.save("template_no_data.xlsx")
    # i=1
    # if i==1: 
        # break


