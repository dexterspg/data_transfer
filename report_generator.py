import pandas as pd
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from id_generator import IdGenerator
from sheet_utils import SheetUtils
from sheet_model import Sheet
from logging_util import LoggingUtil

mandatory_font= Font(color="00FF9B9B")
logger = LoggingUtil.setup_logger('ExcelProcessor')

class ExcelProcessor:

    def __init__(self, input_file, template_file, output_file, config_file, template_header_row=1, input_header_row=1, data_row_start=None) :
        self.input_file =input_file 
        self.output_file =output_file
        self.template_wb= load_workbook(template_file)
        self._load_config(config_file)
        self.template_header_row = template_header_row
        self.input_header_row = input_header_row
        self.data_row_start= self.template_header_row+1 if data_row_start is None else data_row_start
        self.limitRows = None
        self.number_of_rows_to_skip=0

    def _load_config(self, config_file):
        """Load the configuration from JSON file"""
        with open(config_file, 'r') as f:
            self.config = json.load(f)

    def _extract_with_regex(self, text, pattern):
        """Extract value using regex pattern"""
        if pd.isna(text):
            return None
        match = re.search(pattern, str(text))
        if match:
            # If there are capture groups, return the first one
            if match.groups():
                return match.group(1)
            # Otherwise return the entire match
            return match.group(0)
        return None

    def _save_workbook(self):
        self.template_wb.save(self.output_file)
        print(f"Processing complete. Output saved to {self.output_file}")

    def _remove_duplicates(self, df):
        """Remove duplicate rows from DataFrame"""
        return df.drop_duplicates()

    def set_data_row_start(self, data_row_start):
        self.data_row_start = data_row_start

    def set_limit_rows(self, limitRows):
        self.limitRows= limitRows

    def get_mapping_props(self, header, input_df):
        in_header_props=None
        sheet_name = self.config['sheet_name']
        if header in self.config['mappings'].keys():
            in_header_props= self.config['mappings'][header]

            if not in_header_props:
                logger.error(f"Warning: no input mapping found for '{header}' for Sheet '{sheet_name}'") 

            default_val= in_header_props['default']
            output_col = in_header_props['external_column']

            if not default_val and not output_col:
                logger.error(f"Warning: no mapping found for '{header}' for Sheet '{sheet_name}'")
                in_header_props=None
                return in_header_props

            if output_col and output_col not in input_df.columns:
                logger.warning(f"Warning: no input field found for '{header}' for Sheet '{sheet_name}'")
        return in_header_props

    def set_number_of_last_rows_to_drop(self, number_of_rows_to_skip):
        self.number_of_rows_to_skip=number_of_rows_to_skip

    def process(self):
        """Process the input file according to the template and mappings"""
        input_df = pd.read_excel(self.input_file, header =self.input_header_row, nrows=self.limitRows,
                                 skipfooter=self.number_of_rows_to_skip)

        sheet_name = self.config['sheet_name']
        if sheet_name not in self.template_wb.sheetnames:
            logger.error(f"Warning: Sheet '{sheet_name}' not found in template")
            return

        template_sheet = Sheet(self.template_wb[sheet_name], self.template_header_row, self.data_row_start)
        mandatory_fields = self.config['mandatory_fields']

        header_to_autogenerate_id=()
        for header in template_sheet.get_headers():
            isMandatory = header in mandatory_fields
            in_header_props=self.get_mapping_props(header, input_df) 

            if not in_header_props:
                continue

            default_val= in_header_props['default']
            output_col = in_header_props['external_column']
            ext_col_exists = in_header_props['external_column'] != ""

            if in_header_props.get('prefix', "") != "":
                id_generator=IdGenerator(in_header_props['prefix'])
            else:
                id_generator=IdGenerator()

            for r_idx in range(len(input_df)):
                if default_val != "" and ( not ext_col_exists or
                  (output_col and output_col not in input_df.columns)):
                    value=default_val
                    if default_val == "autogenerate":
                        # value=id_generator.generate_id()
                        header_to_autogenerate_id+=(header,)
                else:
                    row_data = {}
                    row = input_df.iloc[r_idx]
                    if 'regex' in in_header_props and in_header_props['regex'] != "":
                        value = self._extract_with_regex(row[output_col], in_header_props['regex'])
                        row_data[header] = value
                    else:
                        value=row[output_col]

                        if isMandatory and not value:
                            if default_val:
                                value=default_val 
                        else:
                            print(f"Warning: No default value set for mandatory field '{header}' for Sheet '{sheet_name}'")

                cell =template_sheet.cell(row=r_idx+self.data_row_start, column=template_sheet.get_col_idx(header), value=value)
                if isMandatory and value == default_val:
                    cell.font = mandatory_font


        logger.info(f"Removing duplicates for colummn {template_sheet.sheet_name()}")
        sheet_no_duplicates=SheetUtils._remove_duplicates_for_sheet(template_sheet, self.data_row_start, self.data_row_start + len(input_df) - 1, (1, len(template_sheet.get_headers())))
        #
        # data_row_range=(sheet_no_duplicates.max_row-template_sheet.header_idx())
        #
        # for header in header_to_autogenerate_id:
        #     in_header_props= self.config['mappings'][header]
        #
        #     if in_header_props.get('prefix', "") != "":
        #         id_generator=IdGenerator(in_header_props['prefix'])
        #     else:
        #         id_generator=IdGenerator()
        #
        #     for r_idx in range(data_row_range):
        #         value=id_generator.generate_id()
        #         cell =template_sheet.cell(row=r_idx+self.data_row_start, column=template_sheet.get_col_idx(header), value=value)
