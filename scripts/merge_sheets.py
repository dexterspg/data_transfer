from openpyxl.reader import excel
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from typing import Tuple, List

def extract_styles(file_path, sheet_name):
    """
    Extract column styles from the header row of the original Excel file.
    Assumes the header is located on row 3.
    Returns a dictionary where keys are header names and values are the number formats.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    column_styles = {}
    # Iterate over columns in header row (row 3)
    for col in ws.iter_cols(min_row=3, max_row=3, min_col=1, max_col=ws.max_column):
        header_value = col[0].value
        if header_value is not None:
            # Store the number format for this header
            column_styles[str(header_value)] = col[0].number_format
    return column_styles

def apply_styles(writer, sheet_name, combined_styles):
    """
    Apply formatting to the merged Excel sheet.
    For each column, if the header is in our combined_styles dictionary, apply that number format.
    Also, if the header contains "Date", force a standard "DD/MM/YYYY" format.
    Sets a default column width.
    """
    wb = writer.book
    ws = writer.sheets[sheet_name]
    
    # Iterate over each column in the sheet.
    # Note: the header is in row 3.
    for col_num in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=3, column=col_num)
        header_value = header_cell.value
        if header_value is None:
            continue  # Skip empty header cells
        header_str = str(header_value)
        
        # Determine the desired number format.
        if "Date" in header_str:
            number_format = "DD/MM/YYYY"
        elif header_str in combined_styles:
            number_format = combined_styles[header_str]
        else:
            number_format = None
        
        # Apply the determined number format to all cells in the data rows (starting from row 4)
        for row_num in range(4, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if number_format:
                cell.number_format = number_format
        
        # Optionally set a default column width (adjust if needed)
        col_letter = ws.cell(row=3, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = 15


def main():
    # Load DataFrames (headers are on row 3)
    excel_file = pd.ExcelFile("input_data.xlsx")
    start_time = time.time()
    # excel_file = pd.read_excel("input_data.xlsx", header=3, sheet_name='S5')
    # merge_file_df=pd.read_excel("merge.xlsx" , header=3, nrows=10)
    # excel_file = excel_file[:-1]
    # print(excel_file)
    # print(merge_file_df.tail())
  
    sheet_names : Tuple[str, ...]  = tuple(excel_file.sheet_names)
    print(sheet_names)

    sheet_dfs : List[pd.DataFrame] = [ excel_file.parse(sheet_name=sheet_name, header=3) for sheet_name in sheet_names]
    # sheet_df = excel_file.parse(sheet_names)


    # for sheet_df in sheet_dfs:
        # sheet_df = sheet_df.iloc[:-1]
        # print(f"max_row {len(sheet_df)}")

    # merge_df_s12 = sheet_dfs[0][:-1].merge(
        # sheet_dfs[1][:-1], on='Property Code 1 - Prim Prop Code', how='left'
    # )

    # merge_df_s24 = merge_df_s12.merge(
        # sheet_dfs[3][:-1], on='Property Code 1 - Prim Prop Code', how='left'
    # )

    # merge_df_sm5 = merge_file_df.merge(
        # excel_file, on='Lease Code 1 - Prim Lease Code', how='left'
    # )

    # merge_df_s56 = merge_df_s45.merge(
        # sheet_dfs[5][:-1], on='Lease Code 1 - Prim Lease Code', how='left'
    # )

    # Write the merged DataFrame to Excel.
    # We use startrow=3 so that the header matches the original layout.
    with pd.ExcelWriter("merge1.xlsx", engine="openpyxl") as writer:
        merge_df_sm5.to_excel(writer, index=False, sheet_name="Merged Data", startrow=3)
        # Reapply styles to the written sheet.
        # apply_styles(writer, "Merged Data", combined_styles)
    
    duration = time.time() - start_time

    print("Data successfully merged and exported with restored formatting!")
    print(f"Time taken: {duration} seconds")

if __name__ == "__main__":
    main()

