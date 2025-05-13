import pandas as pd
from openpyxl import load_workbook



df = pd.read_excel("input_prolease_398.xlsx", header =3, nrows=100)
# df_tem = pd.read_excel("template_init_accounting_on_complete.xlsx", header=2)
# df_out = pd.read_excel("output_nre.xlsx", header = 2)

# print(df)

# print(df_in.head())
# print(df_tem.head())
# print(df_out.head())
# print(df.columns)
# print(df.iat[0,1])

# print(df_index.columns)


# matching_indices = df_in.index[df_in['Name'].str.contains('ABC', na=False].tolist()

# matching_columns = [ (idx, col) for idx, col in enumerate(df.columns) if 'Floor' in col]

matching_columns = [ (idx, col) for idx,col in enumerate(df.columns) if 'Floor' in col]

# print(matching_columns)

# row_values=[]
for col_idx,floor_col in matching_columns:
    row_values=df[floor_col].tolist()
    # print(row_values)
    if 'Office 01' in row_values:
        print(floor_col)
        print(row_values)
        idx =df.index[df[floor_col] =='Office 01'].tolist()
        print(idx)
    print(df.iloc[1, 455])

# print(df.iloc[455,2])

    # if 'Suite' in row_values:
        # print(floor_col)
        # print("\n=====================")
        # print(row_values)

# print(row_values)


    
# for row_idx, row in df_in.iterrows():
    # if df.at[row_idx, ''] ==''



# if 'Floor' in df_in.columns:
    # print(pd.index[])

# print(df['LocationId'])

# wb = load_workbook("input_prolease_398.xlsx")
# i=0
# for sheet_name in wb.sheetnames:
    # sheet = wb[sheet_name]

    # print(sheet.title)

    # headers = [ cell.value for cell in sheet[3]]
    # print(headers)

    # for header in headers:
        # print(header)
        # for row in 
        #
        # for cell in row:
            # print(cell.value)
    # for row in sheet.iter_rows(min_row=4 ,values_only=True):
        # print(row)
    # for cell in sheet[4]:
    #     print(cell.value)
    #     cell.value=None
    #      
# location_sheet= wb['LocationAreaHistory']







# for cell in location_sheet[2]:
    # if cell.value in ['1', 'BU1','S1', 'L1']:
    # if cell.value != None:
        # cell.value =None
        # print(cell.column, cell.value) 
# wb.save("template.xlsx")
    # i=1
    # if i==1: 
        # break


