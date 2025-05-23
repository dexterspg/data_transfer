from dataclasses import dataclass, field
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.cell import Cell
from typing import cast

@dataclass
class ExcelCellStyle:
    font: Font = field(default_factory=Font)
    fill: PatternFill = field(default_factory=PatternFill)
    border: Border = field(default_factory=Border)
    alignment: Alignment = field(default_factory=Alignment)
    number_format: str = "General"

    def __str__(self) -> str:
        fill_rgb = self.fill.fgColor.rgb if self.fill and self.fill.fgColor else 'N/A'
        return (f"Font: {self.font.name}, {self.font.size}, Bold: {self.font.bold}\n"
                f"Fill Color: {fill_rgb}\n"
                f"Border: {self.border}\n"
                f"Alignment: {self.alignment.horizontal}\n"
                f"Number Format: {self.number_format}")

    @classmethod
    def from_cell(cls, cell: Cell) -> 'ExcelCellStyle':
        return cls(
            font=cast(Font, cell.font),
            fill=cast(PatternFill, cell.fill),
            border=cast(Border, cell.border),
            alignment=cast(Alignment, cell.alignment),
            number_format=cell.number_format
        )

