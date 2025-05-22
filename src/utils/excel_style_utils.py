from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


def _apply_date_format(cell, date_format):
    date_style = NamedStyle(name="custom_date_format")
    date_style.number_format = date_format
    print(f"Cell value: {cell.value}")
    # cell = sheet.cell(row=row, column=col)
    cell.style = date_style

def _check_cell_format(sheet, row, col):
    """Check the number format of a specific cell."""
    return sheet.cell(row=row, column=col).number_format

def _apply_date_format_to_df(df):
    pass


