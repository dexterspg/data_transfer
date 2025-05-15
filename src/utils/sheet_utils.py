import pandas as pd
from sheet_model import Sheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment

class SheetUtils:

    @staticmethod
    def _ensure_sheet_instance(sheet):
        if isinstance(sheet, Sheet):
            return sheet.get_sheet() 
        return sheet

    @staticmethod
    def _remove_duplicates_for_sheet(sheet, start_row, end_row, col_range):
        """Remove duplicate rows from the sheet"""
        sheet_obj=SheetUtils._ensure_sheet_instance(sheet)
        seen = set()
        rows_to_delete = []
        for row in sheet_obj.iter_rows(min_row=start_row, max_row=end_row, min_col=col_range[0], max_col=col_range[1]):
            row_values = tuple(str(cell.value) for cell in row)
            if row_values in seen:
                rows_to_delete.append(row[0].row)
            else:
                seen.add(row_values)

        print(len(seen))
        for row_idx in sorted(rows_to_delete, reverse=True):
            sheet_obj.delete_rows(row_idx)
        sheet.set_sheet(sheet_obj)
        return sheet.get_self()

    @staticmethod
    def _remove_duplicates(df):
        """Remove duplicate rows from DataFrame"""
        return df.drop_duplicates()

    @staticmethod
    def _remove_duplicates_for_sheet_use_df(sheet):
        df =sheet.to_data_frame().drop_duplicates().reset_index(drop=True)

        ws=sheet.get_sheet()
        format_map = {}  
        for row_idx in range(sheet.get_data_row_start(), ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                format_map[(row_idx, col_idx)] = {
                    "font": cell.font.copy(),
                    "fill": cell.fill.copy(),
                    "border": cell.border.copy(),
                    "alignment": cell.alignment.copy()
                }
         
        ws.delete_rows(sheet.get_data_row_start(), ws.max_row)

# Ensure new rows are written before restoring formatting
        for row_idx, row_values in enumerate(df.itertuples(index=False, name=None), start=sheet.get_data_row_start()):
            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Restore formatting after writing new values
        for (row_idx, col_idx), styles in format_map.items():
            if row_idx >= sheet.get_data_row_start() and row_idx <= ws.max_row:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = styles["font"]
                cell.fill = styles["fill"]
                cell.border = styles["border"]
                cell.alignment = styles["alignment"]

        sheet.set_sheet(ws)
        return sheet
         

    @staticmethod
    def remove_duplicates_for_single_column(df, column_name):
        return df[[column_name]].drop_duplicates().reset_index(drop=True)

    @staticmethod 
    def clear_row(sheet, min_row):
        for row in sheet.iter_rows(min_row=min_row):
            for cell in row:
                cell.value = None

    @staticmethod
    def _clear_rows(file, min_row):
        wb = load_workbook(file)
        for sheet in wb.worksheets:
            SheetUtils.clear_row(sheet, min_row)


    @staticmethod
    def _num_or_rows(sheet):
        sheet_obj=SheetUtils._ensure_sheet_instance(sheet)

    # @staticmethod
    # def set_values(sheet : Sheet, value):
    #     for row in sheet_obj.iter_rows(min_row=sheet.get_data_row_start()):
    #         for cell in row: 
    #             if cell.value is not None:
    #                 cell.value = value
    #         
    @staticmethod
    def write_dataframe_to_sheet(df, sheet, start_row=2):
        # for col_idx, col_name in enumerate(df.columns, start=1):
            # sheet.cell(row=start_row, column=col_idx, value=col_name)

        # Write data rows starting from the next row
        # for row in df.values.tolist():
            # sheet.append(row)

       # Write data rows (starting from the next row)
        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
