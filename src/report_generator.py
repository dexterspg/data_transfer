import logging
from nre_enums import *
from os import waitpid
import pandas as pd
import json
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font
from id_generator import IdGenerator
from utils import  LoggingUtil
from utils.regex_utils import _extract_with_regex
from sheet_model import Sheet
from rules import _handle_rules_column, _handle_rules_row
from nre_enums import LocationLegalEntityColumns, get_sheet_enum
from create_documents import save_document_indices, retrieve_document_indices

mandatory_font= Font(color="00FF9B9B")
logger = LoggingUtil.setup_logger('ExcelProcessor', console_level=logging.DEBUG)

class ExcelProcessor:

    def __init__(self, input_file, template_file, output_file, config_file, template_header_row=1, input_header_row=1, data_row_start=None) :
        self.input_file =input_file 
        self.output_file =output_file
        self.template_wb= load_workbook(template_file)
        self._load_config(config_file)
        self.template_header_row = template_header_row
        self.input_header_row = input_header_row
        self.data_row_start= self.template_header_row+1 if data_row_start is None else data_row_start
        self.limitRows = None
        self.number_of_rows_to_skip=0

    def _load_config(self, config_file):
        """Load the configuration from JSON file"""
        with open(config_file, 'r') as f:
            self.config = json.load(f)

    def _save_workbook(self):
        self.template_wb.save(self.output_file)
        logger.info(f"Processing complete. Output saved to {self.output_file}")

    def set_data_row_start(self, data_row_start):
        self.data_row_start = data_row_start

    def set_limit_rows(self, limitRows):
        self.limitRows= limitRows

    def valid_header_mapping(self, header : str) -> bool:
        sheet_name = self.config['sheet_name']
        if header in self.config['mappings'].keys():
            in_header_props: dict = self.config['mappings'][header]

            if not in_header_props:
                logger.warning(f"Warning: no input mapping found for '{header}' for Sheet '{sheet_name}'") 

            default_val: str= in_header_props.get('default', None)
            output_col : str = in_header_props.get('external_column',None)

            if not default_val and not output_col:
                logger.error(f"Warning: no mapping found for '{header}' for Sheet '{sheet_name}'")
                return False 

            return True
        return False 


    def autogenerate_cell_ids(self,template_sheet, prefix_dict, data_row_range):
        for header, prefix in prefix_dict.items():
            id_generator=IdGenerator(prefix) if prefix else IdGenerator()
            for r_idx in range(data_row_range):
                value=id_generator.generate_id()
                cell =template_sheet.cell(row=r_idx+self.data_row_start, column=template_sheet.get_col_idx(header), value=value)
                cell.font=mandatory_font
    
    def set_number_of_last_rows_to_drop(self, number_of_rows_to_skip):
        self.number_of_rows_to_skip=number_of_rows_to_skip

    # def populateIdsFromReference(self, header, reference):
        # reference_sheet=self.template_wb[reference]

    def process(self):
        """Process the input file according to the template and mappings"""
        input_df = pd.read_excel(io=self.input_file, header=self.input_header_row, nrows=self.limitRows, skipfooter=self.number_of_rows_to_skip)
        sheet_name = self.config['sheet_name']
        reference : str = self.config.get('has_reference', None)
        if sheet_name not in self.template_wb.sheetnames:
            logger.error(f"Warning: Sheet '{sheet_name}' not found in template")
            return

        template_sheet : Sheet= Sheet(self.template_wb[sheet_name], self.template_header_row, self.data_row_start)
        mandatory_fields = self.config['mandatory_fields']


        header_to_autogenerate_id={}
        original_input_colums = input_df.columns.copy()
        input_df.columns=input_df.columns.str.replace(' ', '_').str.replace("/","_").str.replace(".","_").str.replace('-','_')

        # print(template_sheet.get_headers())
        header_to_autogenerate_id={}
        header_to_apply_column_rule={}
        processed_rows=[{} for _ in range(len(input_df))]  
        for header in template_sheet.get_headers():
            logger.info(f"Processing {header} of sheet {sheet_name}")

            if not self.valid_header_mapping(header):
                continue

            in_header_props: dict= self.config['mappings'][header]

            default_val:str = in_header_props.get('default',"")
            output_col :str = in_header_props.get('external_column',"")
            prefix :str = in_header_props.get('prefix',"")

            if default_val=="autogenerate" and prefix !="":
                header_to_autogenerate_id[header]=prefix 

            normalized_col= output_col.replace(" ", "_").replace("/", "_").replace(".", "_").replace('-', '_')
            is_map_in_input_col : bool=  True if normalized_col and normalized_col in input_df.columns else False
            is_regex_exists : bool='regex' in in_header_props and in_header_props['regex'] != ""

            row_start=template_sheet.get_data_row_start()
            for r_idx, row in enumerate(input_df.itertuples(index=False), start =row_start ):
                value=None
                if is_map_in_input_col:
                    text = getattr(row, normalized_col,None)
                    if text is not None and not pd.isna(text):
                        value =  _extract_with_regex(text, in_header_props['regex']) if is_regex_exists else str(text)

                processed_rows[r_idx-row_start][header] = value

        input_df.columns=original_input_colums
        template_df = pd.DataFrame(processed_rows) 
        logger.info(f"Removing duplicates for colummn {template_sheet.sheet_name()}")
        if reference:
            indices = retrieve_document_indices(str(reference))
            duplicated_mask=template_df.duplicated(keep=False)
            template_df = template_df[~duplicated_mask | template_df.index.isin(indices)]
        else:
            template_df = template_df.drop_duplicates().dropna(how="all")
        # print(template_df)

        template_df_indices= template_df.index.tolist()
        print(template_df_indices)
        save_document_indices(sheet_name, template_df_indices)

        header_rules = self.config.get('has_rules')
        if header_rules:
            print("Before")
            print(template_df)
            for header in header_rules:
                found_rule= _handle_rules_column(input_df, get_sheet_enum(sheet_name), header, self.config['mappings'], template_df_indices)
                if not found_rule.empty:
                    template_df[header] = found_rule[header]

        col_index_map = {
            header: template_sheet.get_col_idx(header)
            for header in template_df.columns
        }

        # get_rule=[]
        for r_idx, row in enumerate(template_df.itertuples(index=False), start=self.data_row_start):
            for header in template_df.columns:
                col_idx=col_index_map[header]
                in_header_props: dict= self.config['mappings'][header]
                default_val : str = in_header_props.get('default',"")
                prefix :str = in_header_props.get('prefix',"")
                isMandatory : bool = header in mandatory_fields
                value=getattr(row, header)
                if value:
                    if default_val != "autogenerate" and prefix != "":
                        value = prefix + value
                    cell = template_sheet.cell(
                        row=r_idx, 
                        column=col_idx, 
                        value=value
                    )
                elif default_val and default_val !=  "autogenerate":
                    value = default_val

                    rules = in_header_props.get('rules')
                    if rules and rules=="row":
                        indices=[]
                        indices.append(template_df_indices[r_idx-self.data_row_start])
                        found_rule= _handle_rules_row(input_df, get_sheet_enum(sheet_name), header, self.config['mappings'],indices )
                        if found_rule:
                            value = found_rule
                        else:
                            value = default_val

                            # get_rule.append(value)

                    cell = template_sheet.cell(
                        row=r_idx, 
                        column=col_idx, 
                        value=value
                    )
                    if isMandatory:
                        cell.font = mandatory_font
                elif isMandatory and not default_val:
                    raise Exception(f"Error: mandatory field '{header}' has missing value")

        # if get_rule:
            # print(get_rule)
        print(template_df)

        self.autogenerate_cell_ids(template_sheet, header_to_autogenerate_id, len(template_df))
        
